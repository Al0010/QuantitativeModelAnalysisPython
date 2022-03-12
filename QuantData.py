import pandas as pd
import pandas_datareader.data as web
import datetime as dt
import math
import matplotlib.pyplot as plt
import mplfinance as mpf
import seaborn as sns
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

pd.set_option('display.max_rows',10000,'display.max_columns',10000)

# Create Dataframes
df14 = pd.DataFrame()
df15 = pd.DataFrame()
df16 = pd.DataFrame()
df17 = pd.DataFrame()
df18 = pd.DataFrame()
df19 = pd.DataFrame()
df20 = pd.DataFrame()
df21 = pd.DataFrame()
df22 = pd.DataFrame()
df23 = pd.DataFrame()

df24 = pd.DataFrame()
#df25 = pd.DataFrame()
#df26 = pd.DataFrame()

# Date
start = dt.datetime(2022,1,1)
end = dt.datetime.now()

# Setting index
corn = "ZC=F"
#oil = "CL=F"
#gas = "NG=F"

# Setting ticker 
agro = "AGRO"
adm = "ADM"
bunge = "BG"
elanco = "ELAN"
darling = "DAR"
ingredion = "INGR"
tilray = "TLRY"
sundial_growers = "SNDL"
bark = "BARK"
alico = "ALCO"

# List asset for heat map
list_asset = ["ZC=F", "AGRO", "ADM", "BG", "ELAN", "DAR", "INGR", "TLRY", "SNDL", "BARK", "ALCO"]

# Download Data from Yahoo Finance
df1 = web.DataReader(agro, 'yahoo', start, end)
df2 = web.DataReader(adm, 'yahoo', start, end)
df3 = web.DataReader(bunge, 'yahoo', start, end)
df4 = web.DataReader(elanco, 'yahoo', start, end)
df5 = web.DataReader(darling, 'yahoo', start, end)
df6 = web.DataReader(ingredion, 'yahoo', start, end)
df7 = web.DataReader(tilray, 'yahoo', start, end)
df8 = web.DataReader(sundial_growers, 'yahoo', start, end)
df9 = web.DataReader(bark, 'yahoo', start, end)
df10 = web.DataReader(alico, 'yahoo', start, end)

# Index Data from Yahoo Finance 
df11 = web.DataReader(corn, 'yahoo', start, end)

# Select "AdjClose"
df14['Rendimenti'] = df1['Adj Close'].pct_change()
df15['Rendimenti'] = df2['Adj Close'].pct_change()
df16['Rendimenti'] = df3['Adj Close'].pct_change()
df17['Rendimenti'] = df4['Adj Close'].pct_change()
df18['Rendimenti'] = df5['Adj Close'].pct_change()
df19['Rendimenti'] = df6['Adj Close'].pct_change()
df20['Rendimenti'] = df7['Adj Close'].pct_change()
df21['Rendimenti'] = df8['Adj Close'].pct_change()
df22['Rendimenti'] = df9['Adj Close'].pct_change()
df23['Rendimenti'] = df10['Adj Close'].pct_change()

df24['Rendimenti'] = df11['Adj Close'].pct_change()

# Delete Na
df14.dropna(inplace = True)
df15.dropna(inplace = True)
df16.dropna(inplace = True)
df17.dropna(inplace = True)
df18.dropna(inplace = True)
df19.dropna(inplace = True)
df20.dropna(inplace = True)
df21.dropna(inplace = True)
df22.dropna(inplace = True)
df23.dropna(inplace = True)
df24.dropna(inplace = True)

# Calculate covariance
cov_agro = df14['Rendimenti'].cov(df24['Rendimenti'])
cov_adm = df15['Rendimenti'].cov(df24['Rendimenti'])
cov_bunge = df16['Rendimenti'].cov(df24['Rendimenti'])
cov_elanco = df17['Rendimenti'].cov(df24['Rendimenti'])
cov_darling = df18['Rendimenti'].cov(df24['Rendimenti'])
cov_ingredion = df19['Rendimenti'].cov(df24['Rendimenti'])
cov_tilray = df20['Rendimenti'].cov(df24['Rendimenti'])
cov_sundial = df21['Rendimenti'].cov(df24['Rendimenti'])
cov_bark = df22['Rendimenti'].cov(df24['Rendimenti'])
cov_alico = df23['Rendimenti'].cov(df24['Rendimenti'])

# Calculate variance
var = df24['Rendimenti'].var()

# Calculate beta 
beta_agro = cov_agro/var
beta_adm = cov_adm/var
beta_bunge = cov_bunge/var
beta_elanco = cov_elanco/var
beta_darling = cov_darling/var
beta_ingredion = cov_ingredion/var
beta_tilray = cov_tilray/var
beta_sundial = cov_sundial/var
beta_bark = cov_bark/var
beta_alico = cov_alico/var

print(beta_agro)
print(beta_adm)
print(beta_bunge)
print(beta_elanco)
print(beta_darling)
print(beta_ingredion)
print(beta_tilray)
print(beta_sundial)
print(beta_bark)
print(beta_alico)

# Calculate Correlation Coefficient 
corr_agro = df14['Rendimenti'].corr(df24['Rendimenti'])
corr_adm = df15['Rendimenti'].corr(df24['Rendimenti'])
corr_bunge = df16['Rendimenti'].corr(df24['Rendimenti'])
corr_elanco = df17['Rendimenti'].corr(df24['Rendimenti'])
corr_darling = df18['Rendimenti'].corr(df24['Rendimenti'])
corr_ingredion = df19['Rendimenti'].corr(df24['Rendimenti'])
corr_tilray = df20['Rendimenti'].corr(df24['Rendimenti'])
corr_sundial = df21['Rendimenti'].corr(df24['Rendimenti'])
corr_bark = df22['Rendimenti'].corr(df24['Rendimenti'])
corr_alico = df23['Rendimenti'].corr(df24['Rendimenti'])

print(corr_agro)
print(corr_adm)
print(corr_bunge)
print(corr_elanco)
print(corr_darling)
print(corr_ingredion)
print(corr_tilray)
print(corr_sundial)
print(corr_bark)
print(corr_alico)

# Calculate Variance 
var_agro = df14['Rendimenti'].var()
var_adm = df15['Rendimenti'].var()
var_bunge = df16['Rendimenti'].var()
var_elanco = df17['Rendimenti'].var()
var_darling = df18['Rendimenti'].var()
var_ingredion = df19['Rendimenti'].var()
var_tilray = df20['Rendimenti'].var()
var_sundial = df21['Rendimenti'].var()
var_bark = df22['Rendimenti'].var()
var_alico = df23['Rendimenti'].var()

# Calculate Standard Deviation 
dev_agro = math.sqrt(var_agro)
dev_adm = math.sqrt(var_adm)
dev_bunge = math.sqrt(var_bunge)
dev_elanco = math.sqrt(var_elanco)
dev_darling = math.sqrt(var_darling)
dev_ingredion = math.sqrt(var_ingredion)
dev_tilray = math.sqrt(var_tilray)
dev_sundial = math.sqrt(var_sundial)
dev_bark = math.sqrt(var_bark)
dev_alico = math.sqrt(var_alico)

print(dev_agro)
print(dev_adm)
print(dev_bunge)
print(dev_elanco)
print(dev_darling)
print(dev_ingredion)
print(dev_tilray)
print(dev_sundial)
print(dev_bark)
print(dev_alico)

data = {
	"Agro": {
		"Beta": beta_agro,
		"Corr": corr_agro,
		"Var": var_agro,
		"DevST": dev_agro
	},
	"Adm": {
		"Beta": beta_adm,
		"Corr": corr_adm,
		"Var": var_adm,
		"DevST": dev_adm
	},
	"Bunge": {
		"Beta": beta_bunge,
		"Corr": corr_bunge,
		"Var": var_bunge,
		"DevST": dev_bunge
	},
	"Darling": {
		"Beta": beta_darling,
		"Corr": corr_darling,
		"Var": var_darling,
		"DevST": dev_darling
	},
	"Ingredion": {
		"Beta": beta_ingredion,
		"Corr": corr_ingredion,
		"Var": var_ingredion,
		"DevST": dev_ingredion
    },
	"Tilray": {
		"Beta": beta_tilray,
		"Corr": corr_tilray,
		"Var": var_tilray,
		"DevST": dev_tilray
    },
	"Sundial": {
		"Beta": beta_sundial,
		"Corr": corr_sundial,
		"Var": var_sundial,
		"DevST": dev_sundial
    },
	"Bark": {
		"Beta": beta_bark,
		"Corr": corr_bark,
		"Var": var_bark,
		"DevST": dev_bark
    },
	"Alico": {
		"Beta": beta_alico,
		"Corr": corr_alico,
		"Var": var_alico,
		"DevST": dev_alico
	}
}

wb = Workbook()
ws = wb.active
ws.title = "Data"

headings = ['Stock'] + list(data['Agro'].keys())
ws.append(headings)

for stock in data:
	excel = list(data[stock].values())
	ws.append([stock] + excel)

wb.save("QuantData.xlsx")

# Create Correlation Heat Map 
metric = "Adj Close"

asset = list_asset
colnames = []

first = True

for ticker in asset:
    data = web.DataReader(f'{ticker}', 'yahoo', start, end)
    if first:
        combined = data[[metric]].copy()
        colnames.append(ticker)
        combined.columns = colnames
        first = False
    else:
        combined = combined.join(data[metric])
        colnames.append(ticker)
        combined.columns = colnames

combined = combined.pct_change().corr(method="pearson")
sns.heatmap(combined, annot = True, cmap = "coolwarm")

plt.show()
